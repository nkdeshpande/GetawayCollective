import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HDR = Font(name='Arial', bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='0C3024')
BODY = Font(name='Arial', size=10)
WRAP = Alignment(wrap_text=True, vertical='top')

wb = openpyxl.Workbook()

# ---------------- OPEN ----------------
ws = wb.active
ws.title = 'Open Blanks'
ws.append(['ID', 'Section', 'What is missing', 'My proposed default',
           'Blocks', 'Owner', 'Priority', 'Approve? (Y / amend)'])

open_rows = [
]
for r in open_rows:
    ws.append(r)

# ---------------- RATIFIED ----------------
rt = wb.create_sheet('Ratified')
rt.append(['ID', 'Section', 'Was', 'Ruling', 'Ratified'])

ratified = [
    ['BLANK-01', 'S4 Category', 'Category Constitution (L1-03)',
     'RATIFIED - One category expressed as facets, not five categories. Defined inside L1-02 Brand Constitution. L1-03 slot marked SUPERSEDED in the artifact catalog.', '30 Jul 2026'],
    ['BLANK-02', 'S5 Vision', 'Vision statement', 'RATIFIED - see L1-01 §5.', '30 Jul 2026'],
    ['BLANK-03', 'S6 Mission', 'Mission statement', 'RATIFIED - see L1-01 §6.', '30 Jul 2026'],
    ['BLANK-04', 'S7 Intent', '10yr / 25yr horizon',
     'RATIFIED - 10yr: India defining platform. 25yr: global operating standard. Excellence before expansion.', '30 Jul 2026'],
    ['BLANK-05', 'S11 TIME', 'TIME domain status',
     'RATIFIED - TIME is the fourth sovereign domain. Tracks entitlements, not operational scheduling.', '30 Jul 2026'],
    ['BLANK-06', 'S15 Layers', 'Four competing layer models',
     'RATIFIED - L1-L12 canonical. Truth Stack nests in L6. Sovereign Stack is L11 topology. L1-L8 Perspectives retired.', '30 Jul 2026'],
    ['BLANK-07', 'S17 Values', 'Enterprise organisational values',
     'RATIFIED - Stewardship Before Ownership, Trust Through Transparency, Excellence Through Intentional Design, Long-Term Thinking.', '30 Jul 2026'],
    ['BLANK-08', 'S24a Governance', 'Governance thresholds',
     'RATIFIED - Ordinary >50% of present. Special >=76% of total. Quorum >=60% of total. Tie = NOT APPROVED. Proxy: written, one meeting, no re-delegation. LLP mandated as default vehicle. Source: EP-01 §3.8.', '30 Jul 2026'],
    ['BLANK-09', 'L1-16 Part II', 'Reserve floor',
     'RATIFIED - NOT NAV-linked. Dual reserve funded from Revenue Base: 2.5% Enterprise Administration Reserve + 2.5% Property Sinking Fund, transferred before partner distributions. Floor = greater of (6 months non-operational fixed obligations) or (Board-approved AAMP minimum). OpCo day-to-day opex EXCLUDED. Per-LLP, non-pooled. Bands: >=120 healthy / 110-119 advisory / 100-109 governance alert / <100 constitutional breach. BREACH RESPONSE (final, see BLANK-09a): AUTOMATIC = broadcast notification + discretionary expenditure deferred. BY RESOLUTION ONLY = suspension of distributions, which needs both an active breach AND an Ordinary Resolution. The 5/15/30-day Board escalation machinery is REPEALED. A reserve breach shall NEVER trigger a capital call. Post-stabilisation capital calls are GROWTH capital only - never operating deficits, routine maintenance, or reserve replenishment.', '30 Jul 2026'],
    ['BLANK-10', 'S24b Accreditation', 'Accreditation expiry and lapse behaviour',
     'RATIFIED - Validity 15 WORKING DAYS (transaction-specific, not standing eligibility). No mandatory constitutional renotification schedule. Lapse behaviour = COMPLETE-THEN-SUSPEND: a commitment formally accepted before expiry completes; one not yet accepted lapses automatically. Post-completion status ACCREDITATION_EXPIRED blocks all new capital actions. Existing ownership unaffected. Governance, voting, distribution and information rights SURVIVE lapse - they attach to ownership, not accreditation. Methodology delegated to Enterprise Compliance Policy, not constitutionalised. Higher local legal standard prevails.', '30 Jul 2026'],
    ['BLANK-11', 'S25a', 'Actor vocabulary',
     'RATIFIED - The Member Law. One actor, two states. Investor pre-commitment, Member post-commitment. Irreversible. Invariant I-08.', '30 Jul 2026'],
    ['BLANK-11a', 'S25a', 'Guest ruling',
     'RATIFIED - DEPRECATED from the investment constitution. A capital table has no guest.', '30 Jul 2026'],
    ['BLANK-12', 'S25b', 'Canonical glossary', 'RATIFIED - 50+ PE terms across five vocabulary families.', '30 Jul 2026'],
    ['BLANK-13', 'S27 Boundaries', 'Commercial boundaries',
     'RATIFIED - Concentration <=10% per property AND <=10% of aggregate portfolio, exceedable only for Board-approved cornerstone investors. Jurisdictions: enforceable title, institutional banking, predictable regulation. Minimum ticket NOT constitutionalised - set per LLP by Investment Committee, disclosed in the Investment Memorandum. Prohibited counterparties: sanctions, failed KYC/AML/source-of-funds, relevant financial-crime convictions, unacceptable reputational or fiduciary risk.', '30 Jul 2026'],
    ['BLANK-13a', 'S27 Platform Equity', 'Does GC hold balance-sheet equity in its own SPVs?',
     'RATIFIED - NO. GOVERNANCE WITHOUT OWNERSHIP. GC shall not subscribe for, acquire, or beneficially own equity in any LLP it establishes, governs, administers or manages. Compensation only via Board-approved disclosed fees (Platform Administration, Asset Management, Origination, Project Development, Brand & Digital Services, Technology Platform, Governance Administration). REVERSES the earlier §27 clause which said the platform retains a strategic ownership position in every collection. Consequence: investment decisions are not co-investment decisions; no independent co-investor review needed. Affiliated-entity transactions remain related-party.', '30 Jul 2026'],
    ['BLANK-14', 'S29', 'Design token conflict',
     'RATIFIED - Design Supremacy Clause. GC-DesignSystem.html v3.0 LOCKED is sole visual truth.', '30 Jul 2026'],
    ['BLANK-15', 'S29b', 'Typography', 'RATIFIED - Outfit, Inter, Space Mono, Playfair Display.', '30 Jul 2026'],
    ['BLANK-16', 'S30 Success', 'Enterprise success metrics', 'RATIFIED - 20 investment-focused KPIs.', '30 Jul 2026'],
    ['BLANK-17', 'S31 Failure', 'Constitutional failure definition',
     'RATIFIED - A Constitutional Failure is a GOVERNANCE event, not an operational incident. Six triggers: CF-01 Invariant Breach, CF-02 Distribution Integrity Failure, CF-03 Fiduciary Breach, CF-04 Governance Failure, CF-05 Regulatory/Legal Failure, CF-06 Asset Stewardship Failure. DECLARATION AUTHORITY AMENDED 30 Jul: interim authority is the GC executive team - COO or CEO - until the Governance & Ethics Committee is constituted, at which point authority transfers automatically without amendment. Where the declaring authority is itself implicated, an Independent Constitutional Review Panel appointed by the Board declares. RECORDED WEAKNESS accepted knowingly: in the interim state the officers most likely to be implicated in CF-03 and CF-04 are the same officers deciding whether a failure occurred; the residual exposure is a failure that is never declared, and nothing in the interim structure detects that. Constituting the Committee is the mitigation and should be time-bound. The declaration cannot be vetoed or ignored. The Board may appeal CLASSIFICATION only, within 15 business days, without suspending remediation; the Panel determination is final. Re-ratification within 90 calendar days. Material failures disclosed to affected LLP Partners.', '30 Jul 2026'],
    ['BLANK-18', 'S32a/32b', 'Amendment procedure',
     'RATIFIED - Proposers: Board, Executive Office, Enterprise Governance Office, Enterprise Architect, any Layer Custodian, any Committee within mandate. LLP Partners request via the Board, cannot table directly. 30 calendar days notice. NO emergency amendment process - urgency goes through the Governance Exception process instead. Pre-tabling written review by Executive Office, Governance Office, Enterprise Architecture and Legal/Compliance - advisory, NO veto. Threshold: Special Resolution >=76%. NO unilateral veto by any individual, office, committee or executive. Cascade Rule ACCEPTED: mandatory Constitutional Impact Assessment before the vote, never after. ENTRENCHED PRINCIPLES require UNANIMOUS approval of all voting LLP Partners plus confirmation that investor rights are not diminished: (1) Fiduciary Primacy (2) Investor Ownership Rights (3) Separation of Ownership/Governance/Service Delivery (4) Related-Party Transparency (5) Equality of Partner Rights within class.', '30 Jul 2026'],
    ['BLANK-19', 'S32c Custodians', 'L1 custodian and successor',
     'RATIFIED - OFFICES, NOT INDIVIDUALS. People change; institutions endure. L1 Custodian = Office of Enterprise Governance, Chief Governance Officer. Successor = whoever the Board appoints to that office; until appointed, the Executive Office discharges the duty. Sign-off authorities likewise by office: Enterprise Architecture Office / Design Authority Office / Chief Compliance & Legal Officer / CEO-Managing Partner. Current holders live in the Governance Register, updatable without constitutional amendment.', '30 Jul 2026'],
    ['BLANK-20', 'S33', 'L2 Business Objects', 'RATIFIED - 25 institutional objects across six domains.', '30 Jul 2026'],
    ['BLANK-13b', 'L1-13 EP-01 §4.10a', 'Related-party materiality threshold',
     'RATIFIED - defaults accepted. Material where ANY of: annual value > Rs 50 lakh; OR > 2% of the SPV annual operating expense; OR term > 36 months; OR creates exclusivity, non-compete or termination-restriction. Below threshold: documented, benchmarked, reported quarterly to Audit & Risk. At or above: Board approval. Above Rs 5 crore: Board approval PLUS Special Resolution. ALWAYS material regardless of value: the Management Agreement, the Brand Participation Agreement, and any amendment to either - because those define Stages 1 and 2 of the waterfall and their effect is structural rather than annual. Load-bearing across the entire GC fee model, since GC holds no equity and therefore reaches the LLPs only through service agreements.', '30 Jul 2026'],
    ['BLANK-24', 'S24a Basis of Voting', 'How are voting rights derived?',
     'RATIFIED - Voting rights are proportional to the percentage EQUITY stake held in the vehicle. Equity-weighted, never per-capita, at any threshold. Three consequences: (1) the 10% concentration cap is also a governance cap, so blocking a Special Resolution requires at least three coordinated holders; (2) the 60% quorum is 60% of EQUITY, not of partners; (3) entrenchment unanimity means 100% of total equity, so a holder of any size can block. Equality within class is entrenched. Voting rights attach to ownership, not accreditation.', '30 Jul 2026'],
    ['A1', 'L1-02 Part VII', 'Brand voice conflict: Warm vs never persuades',
     'RATIFIED 31 Jul 2026 - WARM, CONFIDENT, ASSERTIVE, WITH PLEASANTNESS. The two ratified documents were only in conflict because WARM was being read as SOFT. Warmth is about WHO we speak to - a person owed a plain answer. Persuasion is about WHAT WE WANT from them. A message can be warm and want nothing. Addendum A three principles survive inside the four: Sovereign and Unadorned are what CONFIDENT means in practice, Deterministic is what ASSERTIVE means. Intelligent, Unvarnished and Patient are absorbed; Collaborative retired because it made copy ask questions it did not intend to act on. Governs EVERY string the enterprise emits - there is no second voice for error states. All 25 validation messages rewritten. Enforced by scripts/voice-lint.js over 338 member-facing strings: no apology, no hedging, no blame, no softeners, no exclamation marks, and accessible text must name the field first.', '31 Jul 2026'],
    ['A2', 'Addendum A RISK_COLOUR', 'Risk categories: 8 in design system, 10 in registry',
     'RATIFIED 31 Jul 2026 - ACCEPT ALL. Six of ten rendered grey, which makes a risk register unscannable, the one thing a register exists to be. Four design-only names mapped onto registry synonyms (operational to operator, compliance to regulatory); two reused (reputation purple to counterparty, construction copper to interest_rate - both logged as reuse rather than mapping); two new colours (currency #B8873F amber-gold, technology #5A7D9A slate blue). construction and reputation DROPPED - the registry does not track them and adding them would be a §32a amendment. All ten now distinct, nothing grey. A test caught that the fallback was #6B6B6B steel, identical to LEGAL risk, so an unmapped category impersonated a legal one; fallback moved to steelDim.', '31 Jul 2026'],
    ['A3', 'L1-01 §29 + §29-0', 'Design system version: v3.0 locked vs v4.0 referenced',
     'RATIFIED 31 Jul 2026 - INTEGRATED. §29 amended to name GC.SYSTEM as THREE parts with fixed precedence: Core (canonical) then Addendum A (motion, overlays, notifications, brand, full-bleed) then the Accessibility extension (ground variants). Each later part may only ADD; none may alter a value declared by an earlier one. The clause is now VERSION-AGNOSTIC: it binds to GC.SYSTEM as constituted, not to a number. v4.0 ratified as successor to v3.0 after value-by-value verification showed ZERO DRIFT across 17 colours, 4 typefaces, 10 spacing steps, 2 curves and 4 durations. A version bump changing no value is not a constitutional event; one changing a value is a §32a amendment. New §29-0a makes accessibility part of the system rather than a review of it.', '31 Jul 2026'],
    ['B1', 'L1-01 §29-0a', 'Four contrast variants added to the palette',
     'RATIFIED 31 Jul 2026 - SIGNED OFF. forestLight #228A68 (forest on void was 1.38:1, a dark green on near-black, effectively invisible), copperDeep #8C6635 (copper is the CURRENCY token and was 2.18:1 on paper), confirmDeep #177F43, hazardDeep #BE4915. Each holds its original hue and saturation and moves only lightness to the first value clearing 4.5:1. EVERY ORIGINAL TOKEN KEEPS ITS EXACT VALUE - additive, not a change, so §29 supremacy is intact. ON_GROUND makes the ground-specific choice once so it is not remembered everywhere.', '31 Jul 2026'],
    ['B3', 'lib/processes.ts', 'Process step expiry windows',
     'RATIFIED 31 Jul 2026 - ALIGNED. Accreditation decision 15 working days (from §24b); diligence workstream 180 days; independent valuation 365 days (EP-01 §5.14 annual minimum); IC acquisition approval 90 days. The last was invented and is the tightest: a committee approved the asset AT A PRICE, and after a quarter that is a different decision. Accreditation EVIDENCE (steps A1-A5) holds indefinitely and resumes where it stopped - re-verifying an unchanged passport helps nobody. Only the DECISION expires.', '31 Jul 2026'],
    ['BLANK-23', 'L1-01 §23', 'FINANCIAL invariant ID collision',
     'RATIFIED - MERGED. L1-01 §23 and L1-14 independently assigned the same F-identifiers to different rules, violating §14 Naming Authority. The two sets are merged into ONE canonical table of 18 in L1-01 §23, which is now the sole FINANCIAL numbering. Every rule from both survived; only numbers changed. L1-14 Part IV retains its prose but its numbers are VOID and carry an explicit old-to-new mapping table. Resolved before Wave 2 began, because Wave 2 bakes invariant ids into field definitions.', '30 Jul 2026'],
    ['BLANK-09a', 'L1-16 §2.6a', 'Reserve breach - does distribution suspension survive?',
     'RATIFIED - TWO DISTINCT TESTS. (1) PROSPECTIVE: a distribution that would ITSELF take the reserve below floor is rejected automatically and is NOT voteable - a floor a single payment may cross at will is not a floor. (2) EXISTING BREACH: where the reserve is already below floor, scheduled distributions CONTINUE unless suspended by Ordinary Resolution (>50% of equity present). Automatic on any breach: broadcast to Board, Executive Office and affected LLP Partners, plus deferral of discretionary expenditure. RECORDED TENSION accepted knowingly: partners voting on whether to pause their own distributions are voting on their own cash, and the incentive runs toward continuing to distribute from a vehicle already below its liquidity floor. Limited by the non-voteable prospective test, by the breach being broadcast so the decision is made in the open, and by expenditure deferral operating regardless. Residual exposure: a prolonged sub-floor period sustained by repeated small distributions each individually lawful. Reserve Coverage Ratio (UFR-0385) makes it visible; nothing else prevents it.', '30 Jul 2026'],
    ['BLANK-28', 'L1-16 §1.1a', 'Operating Company ranks ahead of reserves and debt service',
     'RATIFIED - ACCEPT AS ORDERED. Debt is owned by the Investment Vehicle; the Operating Company has nothing to do with it. The operator share is CONSIDERATION FOR SERVICES PERFORMED - the same economic class as the staff and utilities it pays out of that share - not a claim on profit. The Vehicle services its own borrowing, and funds its own reserve provisioning, out of what the property generates AFTER it has paid to be operated. Subordinating the operator to the Vehicle financing would ask a service provider to underwrite a debt it neither incurred nor benefits from. NO constitutional cap on the Stage 1 share: the share is a commercial term that varies by asset, and a fixed ceiling would go stale or force amendments. THE CONTROL ALREADY EXISTS ELSEWHERE - the Management Agreement is an ALWAYS-MATERIAL related-party transaction under EP-01 §4.10a, requiring Board approval to set and Board approval again to amend, regardless of value. The residual starvation risk is real, demonstrated in the test suite, and deliberately governed by the instrument rather than by a constitutional number.', '31 Jul 2026'],
    ['BLANK-25', 'L1-01 §25 Vocabulary', 'Should Consumer join the forbidden list?',
     'RATIFIED - YES. Getaway Collective has Investor and Member, not Consumer. A Member owns; they do not consume. Linter now enforces 16 terms; the codebase was already clean.', '30 Jul 2026'],
    ['BLANK-26', 'WAVE-2-ARCHITECTURE-TARGET', 'Status of GC.SYSTEM v3.0 Sovereign OS scaffold',
     'RATIFIED - MIGRATION TARGET, not current state. C:/gc-app remains authoritative and nothing restructures now; an empty Turborepo would constrain L7/L8 decisions the semantic layer has not yet made. Palette reconciles with zero drift (10/10 verified programmatically). Tier 05 Financial Objects do NOT reconcile - none of Fund/Asset/Deal/Portfolio/Cash Flow/Covenant/Fraction is a canonical L2 name; mapping recorded, with Fraction the consequential one since it carries a retail fractional-ownership framing the constitution rejects. UI state COMMITTED renamed SEALED to avoid colliding with Commitment (UFR-0182). Eight-step migration sequence recorded; step 3 (generate Zod from the UFR rather than transcribing) is the one that keeps E-06 true.', '30 Jul 2026'],
    ['BLANK-27', 'WAVE-2-ARCHITECTURE-TARGET §4', 'GC.SYSTEM vocabulary - canonicalise or exempt?',
     'RATIFIED (delegated) - SPLIT BY DESTINATION, NOT CONTENT. System surface (routes, directories, components, types, schema fields, API contracts) MUST be canonical, because it becomes code and would put the linter permanently at war with the codebase: Steward Passport -> Member Passport, EXP Physical Asset Space -> AST Asset Space, PUB Seduction -> PUB Public Root, Steward Investor -> Investor/Member, Tier 05 names -> the 27 canonical objects. The 8-Pillar brand strategy matrix sits OUTSIDE linter scope as an internal marketing-practice artifact whose pillar names never become tables, routes or fields; forcing it through would yield contortions that serve no engineer. CONDITION on the exemption: the framework stays out of packages/ and apps/. The moment a pillar name becomes a directory, component or field, the canonical rule applies to it.', '30 Jul 2026'],
    ['BLANK-21', 'L1-16 Part III', 'Leverage limit',
     'RATIFIED - NO constitutional LTV, portfolio ratio or DSCR. Rationale: multiple LLPs with different asset classes, lenders and risk profiles; a fixed number would either go stale or force amendments. Limits set per LLP by the Investment Committee, documented in the Investment Memorandum and Financing Plan BEFORE debt is incurred. Once approved the limit is HARD for that LLP; exceeding it requires revised proposal + renewed IC recommendation + Special Resolution where investor risk materially changes. Asset-stage variation permitted and disclosed. Recourse vs non-recourse decided per transaction and disclosed before capital commitment. This is what makes the EP-01 §3.8 leverage trigger operable - previously it referenced limits that did not exist.', '30 Jul 2026'],
    ['BLANK-22', 'L1-16 Parts I & IV', 'Brand Co revenue participation formula',
     'RATIFIED - Six-stage waterfall: (1) Operating Company (2) Brand & Digital Company (3) Enterprise Admin Reserve 2.5% (4) Property Sinking Fund 2.5% (5) Debt Service (6) LLP Partner Distributions. NO preferred return, NO catch-up, NO carried interest - GC holds no equity so has no promote. REVENUE BASE defined constitutionally and REFINED 30 Jul: gross operating receipts less ONLY, in respect of a booking - statutory taxes; booking platform fees; channel/OTA commissions; payment settlement and processing charges; guest refunds and chargebacks. CRITICAL READING: "platform fee" means the BOOKING platform fee, NOT the GC Platform Administration Fee. No fee payable to GC or an affiliated division may be deducted in computing the Revenue Base - doing so would rank GC compensation ahead of both reserves, debt service and every partner distribution, and would shrink the base from which the Brand participation and both reserves are calculated. That is a fiduciary inversion and is prohibited. Brand rate is property-specific, set at underwriting, IC-approved, Board-ratified, disclosed in the Investment Memorandum, fixed for term. Participation is consideration for services - never dilutes ownership, never confers governance or voting rights. KPI-linked with Board remedies up to termination. Annual review. Board fairness benchmark required before approval or renewal.', '30 Jul 2026'],
]
for r in ratified:
    rt.append(r)

# ---------------- format ----------------
for sheet, widths in ((ws, [12, 24, 60, 90, 24, 24, 12, 22]), (rt, [12, 22, 34, 110, 14])):
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for c in sheet[1]:
        c.font = HDR
        c.fill = HDR_FILL
        c.alignment = WRAP
    for row in sheet.iter_rows(min_row=2):
        for c in row:
            c.font = BODY
            c.alignment = WRAP
    sheet.freeze_panes = 'A2'

note = wb.create_sheet('Read Me')
for line in [
    ['GETAWAY COLLECTIVE - L1 BLANKS REGISTER - Rev 3'],
    ['Regenerated 30 Jul 2026'],
    [''],
    ['OPEN: 0   (Waves 1-5 fully ratified)'],
    ['RATIFIED: 37'],
    [''],
    ['Column H on Open Blanks: write Y to accept the default, or write the correction.'],
    [''],
    ['All L1 blanks are ratified. Wave 2 has no open constitutional questions.'],
    [''],
    ['BLANK-09a is a one-line confirmation, not a design question.'],
    [''],
    ['AMENDED THIS SESSION (rulings that replaced earlier rulings):'],
    ['  BLANK-09  breach response simplified - 5/15/30-day escalation repealed'],
    ['  BLANK-17  declaration authority - Governance & Ethics Cmte -> COO/CEO interim'],
    ['  BLANK-22  Revenue Base - booking platform fee and settlement charges added'],
]:
    note.append(line)
note.column_dimensions['A'].width = 100
for row in note.iter_rows():
    for c in row:
        c.font = BODY
note['A1'].font = Font(name='Arial', bold=True, size=12)

wb.save(r'C:\gc-app\L1-BLANKS-REGISTER.xlsx')
print('OK  open=%d  ratified=%d' % (ws.max_row - 1, rt.max_row - 1))
