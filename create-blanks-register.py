#!/usr/bin/env python3
"""L1 Constitution - Blanks Register (rev 2, post-ratification 30 Jul 2026)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════
# SHEET 1 — OPEN BLANKS
# ════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Open Blanks"

HEADERS = ["ID", "Section", "What is missing", "My proposed default",
           "Blocks", "Owner", "Priority", "Approve? (Y / amend)"]

ROWS = [
 # ── HIGH ────────────────────────────────────────────────────────
 ("BLANK-06", "S15 Layers",
  "Four layer models coexist: L1-L12 Enterprise / L1-L8 Perspectives / Truth Stack L1-L3 / "
  "Sovereign Stack L0-L4.",
  "PROPOSED: L1-L12 is the enterprise architecture. Truth Stack nests inside L2-L10 as the "
  "data-authority chain. Sovereign Stack L0-L4 is the implementation topology of L11. "
  "L1-L8 Perspectives is a superseded earlier draft - retire it.",
  "All waves", "Enterprise Architecture", "HIGH"),

 ("BLANK-08", "S24 Governance",
  "Governance thresholds. 76% supermajority referenced in voting.ts; no threshold table exists "
  "(simple majority items, supermajority items, quorum, proxy, tie resolution).",
  "PROPOSED: 76% supermajority for capital events (new issuance, exit, SPV amendment, "
  "economic model version). Simple majority (>50%) for operational items (capex approval, "
  "operator change). Quorum 60% of units. Ties fail. Proxy permitted, one level, written. "
  "LEGAL MUST CONFIRM - this is legally binding.",
  "L5, W5", "Legal + Governance", "HIGH"),

 ("BLANK-02", "S5 Vision",
  "Vision statement. No canonical one-sentence vision exists in the corpus.",
  "PROPOSED: 'A world where owning the extraordinary costs nothing but conviction.' "
  "Drafted from the Stewardship thesis. Yours to replace - this is a founder sentence, "
  "not an architecture one.",
  "W1 exit", "Executive Office", "HIGH"),

 ("BLANK-03", "S6 Mission",
  "Enterprise Mission. Exists per-domain and per-workspace but never at enterprise level.",
  "PROPOSED: 'We structure architectural assets into governed instruments, so that capital "
  "earns yield and silence at the same time.' Yours to replace.",
  "W1 exit", "Executive Office", "HIGH"),

 ("BLANK-07", "S17 Values",
  "Enterprise Values (organisational). Design canon and service canon exist; org values do not.",
  "PROPOSED 4: Precision over speed. Provenance over assertion. Restraint over reach. "
  "Accountability over deniability. Each mirrors an existing invariant so the org behaves "
  "the way the system does.",
  "W1 exit", "Executive Office", "HIGH"),

 ("BLANK-12", "S25 Vocabulary",
  "Glossary completion. 6 forbidden + 16 approved terms now exist; canon calls for 50-100. "
  "Gaps in capital, time and service language.",
  "PROPOSED capital: Commitment (intent) -> Allocation (reserved) -> Subscription (executed). "
  "PROPOSED time: Claim (a Member taking a unit-day), never 'reservation'. "
  "PROPOSED service: Signal (system->Member), Request (Member->system). Confirm the triplets.",
  "W1 exit, W2", "Knowledge Office", "HIGH"),

 ("BLANK-11a", "S25a Member Law",
  "Guest ruling confirmation. I retained Guest as a PRESENCE role (any person in residence "
  "during a Journey, incl. a Member's party), never an identity type, never in FINANCIAL.",
  "PROPOSED: keep Guest as above. Alternative is to deprecate it entirely and say Member "
  "throughout - but then a Member's accompanying party has no noun. Narrow; does not block L2.",
  "L2", "Brand", "HIGH"),

 # ── MEDIUM ──────────────────────────────────────────────────────
 ("BLANK-16", "S30 Success",
  "Enterprise Success Metrics. Seven families named, none populated.",
  "PROPOSED 2 per family. Experience: NPS, silence-interruption count/stay. Financial: net "
  "yield vs model, reserve coverage ratio. Operational: unit-day utilisation, incident MTTR. "
  "Brand: unprompted recall, inbound/outbound lead ratio. Technology: p95 interaction latency "
  "(target 0ms perceived), invariant-violation count (target 0). Sustainability + wellbeing: "
  "NEED YOUR INPUT - no source data.",
  "W10", "Executive Office", "MEDIUM"),

 ("BLANK-10", "S24 Invariants",
  "Accreditation expiry period, renotification schedule, and treatment of in-flight capital "
  "actions when accreditation lapses mid-transaction.",
  "PROPOSED: 12-month validity. Renotify at T-60, T-30, T-7 days. In-flight actions at lapse "
  "are FROZEN, not cancelled - capital is neither taken nor released until re-accreditation "
  "or explicit withdrawal. Confirm the freeze behaviour; it is the contentious part.",
  "L4, W3", "Compliance", "MEDIUM"),

 ("BLANK-09", "S24 Invariants",
  "Reserve shortfall trigger. Canon says shortfalls trigger governance review; no floor defined.",
  "PROPOSED: review compelled when reserve balance falls below 6 months of modelled OpEx, "
  "OR below 1.5% of asset NAV, whichever is higher. Both are placeholders - Finance must set "
  "the real numbers.",
  "L5, W5", "Finance", "MEDIUM"),

 ("BLANK-13", "S26 Boundaries",
  "Commercial boundaries: min ticket, max single-holder concentration, permitted jurisdictions, "
  "prohibited counterparties, whether GC holds balance-sheet equity in its own SPVs.",
  "CANNOT DEFAULT. Every one of these is a legal and strategy decision with no basis in the "
  "source corpus. The single most consequential: does GC take balance-sheet equity in its own "
  "SPVs? That answer changes the governance model and the conflict-of-interest disclosure.",
  "L5, W5", "Executive Office + Legal", "MEDIUM"),

 ("BLANK-01", "S4 Category",
  "Category Constitution (L1-03). Five terms named (Sensory Getaways, Emotional Infrastructure, "
  "Studio Hospitality, Infrastructure Sovereignty, Luxury Operating System) but undefined.",
  "PROPOSED: treat these as ONE category expressed five ways, not five categories. Write a "
  "single L1-03 defining the category once, with the five as facets. Five separate category "
  "definitions would fragment positioning.",
  "W1 exit", "Strategy", "MEDIUM"),

 ("BLANK-04", "S7 Intent",
  "Long-term intent: 10-year and 25-year position. Single-market, geographic expansion, or "
  "license the OS to third-party asset owners?",
  "CANNOT DEFAULT - and this one is load-bearing. If GC ever licenses the operating system, "
  "L2 needs multi-tenancy from Wave 2 and L11 needs per-tenant data residency. Retrofitting "
  "either is a rebuild, not a migration. Answer before W2 even if the horizon is uncertain.",
  "L2, L11", "Executive Office", "MEDIUM"),

 # ── LOW ─────────────────────────────────────────────────────────
 ("BLANK-17", "S31 Failure",
  "Constitutional failure definition. Conditions under which the enterprise declares failure "
  "and re-ratifies.",
  "PROPOSED 3 triggers: (1) an invariant violated in production and shipped, (2) a distribution "
  "executed in error, (3) a stewardship promise broken at scale. Any one compels re-ratification "
  "of the affected layer.",
  "W10", "Governance", "LOW"),

 ("BLANK-18", "S32 Amendment",
  "Amendment procedure: who may propose, notice period, approval threshold, who holds veto.",
  "PROPOSED: any layer steward may propose. 30-day notice. Executive Office + Enterprise "
  "Architecture must both approve. Executive Office holds veto. Downstream re-ratification "
  "enumerated automatically before the vote, never after.",
  "W10", "Governance", "LOW"),

 ("BLANK-19", "S32 Amendment",
  "Stewardship of L1: named custodian and designated successor.",
  "CANNOT DEFAULT - these are two names. Everything else in W10 keys off having them.",
  "W1 exit", "Executive Office", "LOW"),
]

RESOLVED = [
 ("BLANK-06", "S15", "Layer model reconciliation (L1-L12 canonical)",
  "RATIFIED 30 Jul 2026 - The canonical twelve-layer model is: L1 Constitution → L2 Business Objects → "
  "L3 Capabilities → L4 Lifecycles → L5 Governance → L6 Data Authority → L7 Applications → L8 UX → "
  "L9 Analytics → L10 Persistence → L11 Runtime → L12 Observability. Truth Stack nests in L6. "
  "Sovereign Stack is L11 topology. L1-L8 Perspectives archive only.", "You"),
 ("BLANK-05", "S11", "TIME is a fourth sovereign domain",
  "RATIFIED 30 Jul 2026 - TIME is a first-class sovereign domain (alongside ASSET, IDENTITY, FINANCIAL). "
  "TIME tracks the scarcity of nights that an Investor receives as part of their Ownership Interest. "
  "TIME represents entitlements, not operational scheduling. Operational scheduling (real-time blocking, "
  "availability management) occurs in L4 and L5, not at domain level. SYSTEM is not a domain — it is L11.", "You"),
 ("BLANK-20", "S33", "L2 Business Objects (18-object institutional model)",
  "RATIFIED 30 Jul 2026 - MAJOR CORRECTION: Getaway Collective is a PE platform, not an operating company. "
  "The canonical L2 BOs are 18 objects organized by domain: Enterprise (Organization, Fund, SPV, Portfolio) · "
  "Assets (Property, Acquisition, Disposition, Valuation) · Capital (Offering, Commitment, Capital Call, Investment, "
  "Ownership Position, Distribution) · Investors (Investor, Suitability, Accreditation, Relationship) · "
  "Governance (Agreement, Resolution, Policy, Committee, Vote) · Performance (KPI, Benchmark, Report, Forecast) · "
  "Risk (Risk Register, Compliance, Audit) · Partners (Operating Partner, Advisor, Developer, Lender) · "
  "Knowledge (Investment Thesis, Market Intelligence, Research, Due Diligence). Removed: Studio, Journey, Experience, "
  "Service, Guest (operating company concerns). Constitutional boundary: GC owns/invests; Sensory Getaways operates.", "You"),
 ("BLANK-02", "S5", "Vision statement",
  "RATIFIED 30 Jul 2026 - 'To create the world's most trusted collective ownership platform "
  "for extraordinary places, where architecture, stewardship, and technology transform real "
  "estate into a lifelong source of restoration, belonging, and enduring value.' With constitutional "
  "expansion defining the vision of collective stewardship across generations.", "You"),
 ("BLANK-03", "S6", "Mission statement",
  "RATIFIED 30 Jul 2026 - 'To make exceptional real estate accessible through intelligent "
  "collective ownership by combining carefully curated properties, institutional governance, and "
  "a technology platform that delivers effortless stewardship, transparent operations, and "
  "extraordinary experiences.' With six integration pillars.", "You"),
 ("BLANK-04", "S7", "Long-term horizon (10yr / 25yr)",
  "RATIFIED 30 Jul 2026 - 10-year position: India's defining platform. 25-year position: "
  "global operating standard. Growth by direct ownership, partnerships, regional operators, "
  "technology licensing, governance frameworks. Permanent principle: excellence before expansion. "
  "Strategic doctrine: Years 0–10 prove model; Years 10–25 export operating system.", "You"),
 ("BLANK-07", "S17", "Enterprise organizational values",
  "RATIFIED 30 Jul 2026 - Four values mirroring platform invariants: (1) Stewardship Before "
  "Ownership — permanent stewardship, temporary ownership. (2) Trust Through Transparency — "
  "visibility on decisions, money, performance, governance. (3) Excellence Through Intentional "
  "Design — every interaction deliberate, nothing accidental. (4) Long-Term Thinking — optimize "
  "for decades, durability over speed.", "You"),
 ("BLANK-12", "S25b", "Canonical glossary (50+ terms)",
  "RATIFIED 30 Jul 2026 - Five vocabulary families defined: Capital (Investment, Ownership "
  "Interest, Collective Portfolio, Liquidity Event, Ownership Return, Exceptional Property, "
  "Long-term Value Creation). Time (Reservation, Residency, Allocation, Access Window, Ownership "
  "Calendar, Retreat, Time Allocation). Service (Member presence role, Owner, Concierge, "
  "Stewardship, Property Care, Local Steward, Collective Services, Member Experience). Property "
  "(Property, Residence, Retreat, Collection, Property Profile, Portfolio, Residence, Residence). "
  "Governance (Governance, Collective Standards, Stewardship Contributions, Governance Council, "
  "Platform Steward, Membership Interest).", "You"),
 ("BLANK-13", "S26-27", "Commercial boundaries (minticket, concentration, jurisdiction, GC equity)",
  "RATIFIED 30 Jul 2026 - Investment: minimum institutional viability, no fractions below thresholds. "
  "Concentration: ≤10% per property unless approved. Jurisdictions: enforceable property rights, "
  "transparent titles, institutional banking, predictable regulation. Liquidity: governance-approved "
  "transfers only, no speculation. Platform Equity: strategic position in every collection aligns "
  "incentives without compromising neutrality. Asset Standards: architecture, location, longevity, "
  "operational feasibility, financial sustainability. Capital Discipline: growth scales with maturity.", "You"),
 ("BLANK-16", "S30", "Enterprise success metrics (7 families, 30+ measures)",
  "RATIFIED 30 Jul 2026 - Experience: NPS 75+, zero-interruption target. Financial: yield vs "
  "model ±2%, 6-month reserve. Operational: 65–75% utilization, <4hr MTTR. Brand: 40% unprompted "
  "recall, 3:1 lead ratio. Technology: 0ms P95 latency, zero invariant violations. Sustainability: "
  "carbon tracking, renewable energy %, biodiversity, resource circularity. Wellbeing: Member "
  "return intention, stress reduction, digital detox; Team retention, training, promotion; Community "
  "employment, local procurement, economic impact.", "You"),
 ("BLANK-11a", "S25a", "Guest ruling: deprecate from investment constitution",
  "RATIFIED 30 Jul 2026 - DEPRECATED. A shareholder register has no guest. A capital table has no guest. "
  "A property fund has no guest. If the operator (Sensory Getaways) tracks guests for operational purposes, "
  "that is an operating company concern. The investment platform sees only: Investor → Commitment → Investment "
  "→ Ownership Position → Distribution.", "You"),
 ("BLANK-11", "S25a", "Actor vocabulary: Member / Investor / Steward",
  "RATIFIED 30 Jul 2026 - The Member Law. ONE actor. Investor = pre-commitment state, "
  "Member = post-commitment state. State change on one identity, never a second record. "
  "Irreversible. Steward deprecated as actor noun (philosophy only). User + Customer forbidden. "
  "New invariant I-08 Single Actor.", "You"),
 ("BLANK-14", "S29a", "Design token / palette conflict",
  "RATIFIED 30 Jul 2026 - Design Supremacy Clause. GC-DesignSystem.html (v3.0 LOCKED) is the "
  "sole source of visual truth and overrides every other document without exception. "
  "Bone / Vantablack / Lichen / Cinnabar DEPRECATED. paper = #F2F2F2; #F4F4F0 is void.", "You"),
 ("BLANK-15", "S29b", "Typography reconciliation",
  "RATIFIED 30 Jul 2026 - all four roles stand: Outfit (display), Inter (body), Space Mono "
  "(data/metadata), Playfair Display (editorial, italic only). The two-role reading in "
  "The Architecture of Silence is incomplete, not canonical.", "You"),
]

# ── styling ──────────────────────────────────────────────────────
PRI = {
    "BLOCKER": ("FF3B30", "FFFFFF"),
    "HIGH":    ("E8672E", "FFFFFF"),
    "MEDIUM":  ("C79F6B", "0A0A0A"),
    "LOW":     ("E8E8E6", "0A0A0A"),
}
thin = Side(style="thin", color="9A9A9A")
hf = PatternFill("solid", start_color="0C3024", end_color="0C3024")


def head(sheet, headers):
    sheet.append(headers)
    for c in sheet[1]:
        c.fill = hf
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"


head(ws, HEADERS)
for i, r in enumerate(ROWS, start=2):
    ws.append(list(r) + [""])
    pc = ws.cell(row=i, column=7)
    bg, fg = PRI[r[6]]
    pc.fill = PatternFill("solid", start_color=bg, end_color=bg)
    pc.font = Font(name="Arial", size=10, bold=True, color=fg)
    pc.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=i, column=1).font = Font(name="Arial", size=10, bold=True)
    for col in range(1, 9):
        cell = ws.cell(row=i, column=col)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if col == 7:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=i, column=8).fill = PatternFill("solid", start_color="FFFFCC",
                                                end_color="FFFFCC")
    ws.row_dimensions[i].height = 76

for k, v in {"A": 12, "B": 16, "C": 52, "D": 62, "E": 14,
             "F": 24, "G": 11, "H": 20}.items():
    ws.column_dimensions[k].width = v

n = len(ROWS) + 3
ws.cell(row=n, column=1, value="OPEN").font = Font(name="Arial", size=11, bold=True)
ws.cell(row=n, column=2, value=len(ROWS)).font = Font(name="Arial", size=11, bold=True)
last = len(ROWS) + 1
for off, (lbl, col) in enumerate(
        [("BLOCKER", "FF3B30"), ("HIGH", "E8672E"),
         ("MEDIUM", "C79F6B"), ("LOW", "6B6B6B")], start=1):
    ws.cell(row=n + off, column=1, value=lbl).font = Font(
        name="Arial", size=10, bold=True, color=col)
    ws.cell(row=n + off, column=2,
            value=f'=COUNTIF(G2:G{last},"{lbl}")').font = Font(
        name="Arial", size=10, bold=True)

ws.cell(row=n + 6, column=1,
        value="Column D is MY PROPOSAL, not your instruction. Column H: write Y to accept, "
              "or write the correction. Items marked CANNOT DEFAULT need a real answer.")
ws.cell(row=n + 6, column=1).font = Font(name="Arial", size=10, italic=True)

# ════════════════════════════════════════════════════════════════════
# SHEET 2 — RATIFIED
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Ratified")
head(ws2, ["ID", "Section", "Was", "Ruling", "Ratified by"])
for i, r in enumerate(RESOLVED, start=2):
    ws2.append(list(r))
    ws2.cell(row=i, column=1).font = Font(name="Arial", size=10, bold=True)
    for col in range(1, 6):
        c = ws2.cell(row=i, column=col)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.cell(row=i, column=4).fill = PatternFill("solid", start_color="D9F2E3",
                                                 end_color="D9F2E3")
    ws2.row_dimensions[i].height = 92
for k, v in {"A": 12, "B": 12, "C": 40, "D": 78, "E": 12}.items():
    ws2.column_dimensions[k].width = v

wb.save(r"C:\gc-app\L1-BLANKS-REGISTER.xlsx")
print(f"[OK] rev 2 written - {len(ROWS)} open, {len(RESOLVED)} ratified")
