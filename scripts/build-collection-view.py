"""
GC COLLECTION — FULL VIEW
Builds outputs/GC-COLLECTION-FULL-VIEW.xlsx from the codebase registries.

Source of truth is constants/vehicles.ts and constants/spatial.ts, dumped to
_gc.json by _dump.ts. Nothing here is retyped: if the code changes, rerun and
the workbook changes with it. That is the whole point — a spreadsheet
transcribed by hand becomes a second, quietly diverging canon.

Run:  node --experimental-strip-types ./_dump.ts > ./_gc.json
      python scripts/build-collection-view.py
      python <skill>/scripts/recalc.py outputs/GC-COLLECTION-FULL-VIEW.xlsx
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

D = json.load(open("_gc.json", encoding="utf-8"))
V = D["vehicles"]; E = D["estates"]; C = D["conflicts"]; S = D["stages"]

FONT = "Arial"
H1   = Font(name=FONT, size=16, bold=True, color="1A1A1A")
H2   = Font(name=FONT, size=11, bold=True, color="FFFFFF")
LBL  = Font(name=FONT, size=10, bold=True)
TXT  = Font(name=FONT, size=10)
DIM  = Font(name=FONT, size=9, color="666666")
IN   = Font(name=FONT, size=10, color="0000FF")   # hardcoded input
CALC = Font(name=FONT, size=10, color="000000")   # formula
LINK = Font(name=FONT, size=10, color="008000")   # cross-sheet
BAD  = Font(name=FONT, size=10, bold=True, color="C00000")

HDR  = PatternFill("solid", fgColor="1F3B34")
BAND = PatternFill("solid", fgColor="F2F5F4")
WARN = PatternFill("solid", fgColor="FFF2CC")
STOP = PatternFill("solid", fgColor="FCE4E4")
OK   = PatternFill("solid", fgColor="E6F4EA")

THIN = Side(style="thin", color="D9D9D9")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RUP  = '₹#,##0;(₹#,##0);-'
PCT  = '0.00%'
NUM  = '#,##0;(#,##0);-'

wb = Workbook()


def head(ws, title, subtitle, width):
    ws["A1"] = title; ws["A1"].font = H1
    ws["A2"] = subtitle; ws["A2"].font = DIM
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws.freeze_panes = "A5"


def row(ws, r, values, font=TXT, fill=None, fmt=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = font; c.border = BOX
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if fill: c.fill = fill
        if fmt and isinstance(v, (int, float)): c.number_format = fmt
    return r + 1


def header_row(ws, r, labels, widths=None):
    for i, v in enumerate(labels, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = H2; c.fill = HDR; c.border = BOX
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 30
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return r + 1


# ───────────────────────── 1 · READ ME ─────────────────────────
ws = wb.active; ws.title = "Read Me"
head(ws, "GC COLLECTION — FULL VIEW",
     "Generated from constants/vehicles.ts and constants/spatial.ts. Do not edit: rerun the generator.", 4)
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 96

r = 5
for k, v in [
    ("What this is",
     "Every fact the platform holds about the three vehicles, in one place, folded out of the code "
     "rather than retyped. Rerun the generator and this file changes with it."),
    ("Sources",
     "GC-LLP-INTAKE-TEMPLATE.xlsx (saved 4 Aug 2026) for the LLPs · GETAWAYS SPATIAL LEDGER.xlsx "
     "for the estates, design system and deployment programme."),
    ("Read the Conflicts tab first",
     "Nine contradictions across the two documents and the existing canon. Seven are unresolved and "
     "five of those are blocking, which means no vehicle currently clears the gate to appear on a "
     "public surface. Every figure elsewhere in this workbook should be read against that tab."),
    ("Absent is not zero",
     "A blank cell means the source states nothing. It is never a zero — a reserve floor of zero "
     "reads as a vehicle with no floor rather than one nobody has set."),
    ("Nothing here is appraised",
     "No building exists on any of the three. Every valuation is a project cost and every yield is "
     "a forecast from a model, not an observation."),
    ("Colour",
     "Blue = a figure taken from a source document. Black = computed by a formula in this workbook. "
     "Green = carried from another tab. Red = a conflict or a breach."),
]:
    ws.cell(row=r, column=1, value=k).font = LBL
    ws.cell(row=r, column=1).alignment = Alignment(vertical="top")
    c = ws.cell(row=r, column=2, value=v); c.font = TXT
    c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 46
    r += 1

r += 1
ws.cell(row=r, column=1, value="Tabs").font = LBL
r += 1
for name, what in [
    ("Collection", "One row per vehicle — the headline view."),
    ("Capital", "Stack, offering and unit economics, with the arithmetic live."),
    ("Waterfall", "The six stages, in order, with the 100% check."),
    ("Time & Governance", "Entitlement pool and voting thresholds."),
    ("Estate & Design", "Ground, climate pack, key typologies and footprint."),
    ("Pipeline", "The ten-stage vehicle lifecycle and the construction programme."),
    ("Conflicts", "The nine, with severity and who can settle each."),
]:
    ws.cell(row=r, column=1, value=name).font = LINK
    ws.cell(row=r, column=2, value=what).font = TXT
    r += 1


# ───────────────────────── 2 · COLLECTION ─────────────────────────
ws = wb.create_sheet("Collection")
head(ws, "THE COLLECTION", "One row per vehicle. Blue is stated in a source; black is computed here.", 14)
r = header_row(ws, 4, [
    "Key", "Property", "Asset code", "Registered name", "LLPIN", "Jurisdiction", "Estate",
    "Land area", "Keys", "Vehicle state", "Property state", "Project total (₹)",
    "Units avail.", "Publishable?",
], [12, 22, 11, 24, 12, 30, 16, 30, 7, 13, 14, 16, 11, 13])

first = r
for v in V:
    est = next((e for e in E if e["id"] == v["estate"]), None)
    pub = v["publishable"]
    r = row(ws, r, [
        v["key"], v["propertyName"], v["assetCode"], v["registeredName"],
        v["llpin"] or "not incorporated", v["jurisdiction"],
        est["name"] if est else "not joined",
        v["landArea"], v["keys"], v["lifecycle"], v["propertyLifecycle"],
        v["stack"]["projectTotal"], v["offering"]["available"],
        "YES" if pub["ok"] else "NO",
    ], font=IN, fmt=RUP)
    ws.cell(row=r - 1, column=12).number_format = RUP
    gate = ws.cell(row=r - 1, column=14)
    gate.fill = OK if pub["ok"] else STOP
    gate.font = CALC if pub["ok"] else BAD
    if not pub["ok"]:
        gate.comment = None
        ws.cell(row=r - 1, column=14).value = "NO — " + "; ".join(
            b.split(":")[0] for b in pub["because"])

r += 1
ws.cell(row=r, column=1, value="TOTAL").font = LBL
t = ws.cell(row=r, column=12, value=f"=SUM(L{first}:L{first + len(V) - 1})")
t.font = CALC; t.number_format = RUP; t.border = BOX
k = ws.cell(row=r, column=9, value=f"=SUM(I{first}:I{first + len(V) - 1})")
k.font = CALC; k.number_format = NUM; k.border = BOX

r += 2
ws.cell(row=r, column=1,
        value="Publishable is a gate, not a preference: constants/vehicles.ts publishable() refuses a "
              "public surface to any vehicle carrying a blocking conflict. See the Conflicts tab."
        ).font = DIM
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)


# ───────────────────────── 3 · CAPITAL ─────────────────────────
ws = wb.create_sheet("Capital")
head(ws, "CAPITAL STACK AND OFFERING",
     "Blue from the intake. Black computed here — the checks are live, so a changed input re-tests itself.", 5)
ws.column_dimensions["A"].width = 34
for i in range(2, 6):
    ws.column_dimensions[get_column_letter(i)].width = 21

r = header_row(ws, 4, ["", *[v["propertyName"] for v in V], "Note"])

def cap_row(r, label, path, fmt=RUP, note=""):
    ws.cell(row=r, column=1, value=label).font = LBL
    ws.cell(row=r, column=1).border = BOX
    for i, v in enumerate(V, start=2):
        val = v
        for p in path: val = val[p]
        c = ws.cell(row=r, column=i, value=val)
        c.font = IN; c.number_format = fmt; c.border = BOX
    n = ws.cell(row=r, column=5, value=note); n.font = DIM; n.border = BOX
    n.alignment = Alignment(wrap_text=True, vertical="top")
    return r + 1

start = r
r = cap_row(r, "Land (₹)", ["stack", "land"])
r = cap_row(r, "Formation (₹)", ["stack", "formation"])
eq_check = r
ws.cell(row=r, column=1, value="Equity layer — stack (₹)").font = LBL
ws.cell(row=r, column=1).border = BOX
for i in range(2, 5):
    c = ws.cell(row=r, column=i, value=f"=SUM({get_column_letter(i)}{start}:{get_column_letter(i)}{start+1})")
    c.font = CALC; c.number_format = RUP; c.border = BOX
ws.cell(row=r, column=5, value="Land + formation. Computed, so it cannot drift from its parts.").font = DIM
r += 1

r = cap_row(r, "Equity layer — as stated (₹)", ["stack", "equityLayer"],
            note="Sheet 3 of the intake.")
stated_eq = r - 1
r = cap_row(r, "Equity — offering sheet (₹)", ["offering", "totalEquity"],
            note="Sheet 4. Differs from sheet 3 for Solace — conflict C-03.")
off_eq = r - 1

ws.cell(row=r, column=1, value="Sheets agree?").font = LBL
ws.cell(row=r, column=1).border = BOX
for i in range(2, 5):
    L = get_column_letter(i)
    c = ws.cell(row=r, column=i, value=f'=IF({L}{stated_eq}={L}{off_eq},"agrees","DISAGREES by "&TEXT(ABS({L}{off_eq}-{L}{stated_eq}),"₹#,##0"))')
    c.font = CALC; c.border = BOX
ws.cell(row=r, column=5, value="Live check. Solace fails it by ₹50 lakh.").font = DIM
r += 2

r = cap_row(r, "Facility / debt (₹)", ["stack", "facility"])
r = cap_row(r, "Project total (₹)", ["stack", "projectTotal"])
r += 1

ws.cell(row=r, column=1, value="THE OFFERING").font = LBL
ws.cell(row=r, column=1).fill = BAND
r += 1
prom = r
r = cap_row(r, "Promoter / sponsor (₹)", ["offering", "promoter"],
            note="The sponsor's own money. The modelled canon has no concept of this — conflict C-06.")
offered = r
r = cap_row(r, "Offered to partners (₹)", ["offering", "offered"])

ws.cell(row=r, column=1, value="Promoter share of equity").font = LBL
ws.cell(row=r, column=1).border = BOX
for i in range(2, 5):
    L = get_column_letter(i)
    c = ws.cell(row=r, column=i, value=f"=IF({L}{off_eq}=0,\"\",{L}{prom}/{L}{off_eq})")
    c.font = CALC; c.number_format = PCT; c.border = BOX
ws.cell(row=r, column=5, value="Guarded denominator.").font = DIM
r += 1

units = r
r = cap_row(r, "Units in the offering", ["offering", "units"], fmt=NUM)
price = r
r = cap_row(r, "Unit price (₹)", ["offering", "unitPrice"])

ws.cell(row=r, column=1, value="Units × price = offered?").font = LBL
ws.cell(row=r, column=1).border = BOX
for i in range(2, 5):
    L = get_column_letter(i)
    c = ws.cell(row=r, column=i, value=f'=IF({L}{units}*{L}{price}={L}{offered},"balances","OUT BY "&TEXT({L}{units}*{L}{price}-{L}{offered},"₹#,##0"))')
    c.font = CALC; c.border = BOX
r += 1

sub = r
r = cap_row(r, "Units subscribed", ["offering", "subscribed"], fmt=NUM)
r = cap_row(r, "Units available", ["offering", "available"], fmt=NUM)
r = cap_row(r, "Deposit (₹)", ["offering", "deposit"])
r += 1

ws.cell(row=r, column=1, value="Raised so far (₹)").font = LBL
ws.cell(row=r, column=1).border = BOX
for i in range(2, 5):
    L = get_column_letter(i)
    c = ws.cell(row=r, column=i, value=f"={L}{sub}*{L}{price}")
    c.font = CALC; c.number_format = RUP; c.border = BOX
r += 1
ws.cell(row=r, column=1, value="Still to raise (₹)").font = LBL
ws.cell(row=r, column=1).border = BOX
for i in range(2, 5):
    L = get_column_letter(i)
    c = ws.cell(row=r, column=i, value=f"={L}{offered}-{L}{sub}*{L}{price}")
    c.font = CALC; c.number_format = RUP; c.border = BOX
r += 2

for i, v in enumerate(V, start=2):
    ws.cell(row=r, column=i, value=v["stack"]["covenant"]).font = DIM
ws.cell(row=r, column=1, value="Covenant").font = LBL
r += 1
for i, v in enumerate(V, start=2):
    ws.cell(row=r, column=i, value=v["offering"]["lockIn"]).font = DIM
ws.cell(row=r, column=1, value="Lock-in").font = LBL


# ───────────────────────── 4 · WATERFALL ─────────────────────────
ws = wb.create_sheet("Waterfall")
head(ws, "THE SIX-STAGE WATERFALL",
     "Basis points of gross revenue, in the order they are paid. Debt service is its own stage — folded "
     "inside the partner share it would show a number a quarter of gross never reaches.", 5)
ws.column_dimensions["A"].width = 30
for i in range(2, 6):
    ws.column_dimensions[get_column_letter(i)].width = 21

r = header_row(ws, 4, ["", *[v["propertyName"] for v in V], "Note"])

gross = r
ws.cell(row=r, column=1, value="Gross revenue (₹)").font = LBL
ws.cell(row=r, column=1).border = BOX
for i, v in enumerate(V, start=2):
    c = ws.cell(row=r, column=i, value=v["operating"]["grossRevenue"])
    c.font = IN; c.number_format = RUP; c.border = BOX
r += 1
ws.cell(row=r, column=1, value="ADR (₹) · occupancy").font = LBL
for i, v in enumerate(V, start=2):
    c = ws.cell(row=r, column=i,
                value=f"₹{v['operating']['adr']:,.0f} · {v['operating']['occupancyBps']/100:.0f}%")
    c.font = IN; c.border = BOX
r += 2

STAGES = [("1 Operator", "operator"), ("2 Brand", "brand"), ("3 Admin reserve", "adminReserve"),
          ("4 Sinking fund", "sinkingFund"), ("5 Debt service", "debtService"),
          ("6 To partners", "toPartners")]
bps_start = r
for label, k in STAGES:
    ws.cell(row=r, column=1, value=f"{label} (bps)").font = LBL
    ws.cell(row=r, column=1).border = BOX
    for i, v in enumerate(V, start=2):
        w = v["operating"]["waterfall"]
        c = ws.cell(row=r, column=i, value=(w[k] if w else None))
        c.font = IN; c.number_format = NUM; c.border = BOX
        if not w: c.fill = WARN
    r += 1
bps_end = r - 1

ws.cell(row=r, column=1, value="Sums to 10,000?").font = LBL
ws.cell(row=r, column=1).border = BOX
for i in range(2, 5):
    L = get_column_letter(i)
    c = ws.cell(row=r, column=i,
                value=f'=IF(COUNT({L}{bps_start}:{L}{bps_end})<6,"NOT STATED",'
                      f'IF(SUM({L}{bps_start}:{L}{bps_end})=10000,"closes at 100%",'
                      f'"OUT BY "&SUM({L}{bps_start}:{L}{bps_end})-10000&" bps"))')
    c.font = CALC; c.border = BOX
ws.cell(row=r, column=5,
        value="Solace states no waterfall. Its intake row would read as 100% of gross to partners with no "
              "operator, reserve or debt service — a blank row, not a split. Conflict C-05.").font = DIM
r += 2

ws.cell(row=r, column=1, value="To partners (₹)").font = LBL
ws.cell(row=r, column=1).border = BOX
part = r
for i in range(2, 5):
    L = get_column_letter(i)
    c = ws.cell(row=r, column=i, value=f"=IF({L}{bps_end}=\"\",\"\",{L}{gross}*{L}{bps_end}/10000)")
    c.font = CALC; c.number_format = RUP; c.border = BOX
r += 1

ws.cell(row=r, column=1, value="Forecast yield on equity").font = LBL
ws.cell(row=r, column=1).border = BOX
for i in range(2, 5):
    L = get_column_letter(i)
    c = ws.cell(row=r, column=i, value=f"=IF(OR({L}{part}=\"\",Capital!{L}9=0),\"\",{L}{part}/Capital!{L}9)")
    c.font = LINK; c.number_format = PCT; c.border = BOX
ws.cell(row=r, column=5,
        value="Against the offering sheet's equity. A forecast from a model on an asset that does not "
              "exist — not an observation, and not a promise.").font = DIM
r += 2

ws.cell(row=r, column=1, value="Reserve floor (₹)").font = LBL
ws.cell(row=r, column=1).border = BOX
for i, v in enumerate(V, start=2):
    rf = v["operating"]["reserveFloor"]
    c = ws.cell(row=r, column=i, value=rf)
    c.font = IN; c.number_format = RUP; c.border = BOX
    if rf is None: c.fill = WARN
ws.cell(row=r, column=5, value="Stage 6 does not run at all if paying it would take the reserve below this.").font = DIM


# ───────────────────────── 5 · TIME & GOVERNANCE ─────────────────────────
ws = wb.create_sheet("Time & Governance")
head(ws, "ENTITLEMENT AND GOVERNANCE",
     "Time is an incident of a capital position, not a hospitality product. Voting is contribution-weighted, "
     "and a tie is not approval.", 5)
ws.column_dimensions["A"].width = 32
for i in range(2, 6):
    ws.column_dimensions[get_column_letter(i)].width = 24

r = header_row(ws, 4, ["", *[v["propertyName"] for v in V], "Note"])

def tg(r, label, get, note="", fmt=None):
    ws.cell(row=r, column=1, value=label).font = LBL
    ws.cell(row=r, column=1).border = BOX
    for i, v in enumerate(V, start=2):
        val = get(v)
        c = ws.cell(row=r, column=i, value=val)
        c.font = IN; c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if fmt and isinstance(val, (int, float)): c.number_format = fmt
        if val is None: c.fill = WARN
    n = ws.cell(row=r, column=5, value=note); n.font = DIM; n.border = BOX
    n.alignment = Alignment(wrap_text=True, vertical="top")
    return r + 1

r = tg(r, "Night pool — minimum", lambda v: v["entitlement"]["nightPoolMin"] if v["entitlement"] else None, fmt=NUM)
r = tg(r, "Night pool — maximum", lambda v: v["entitlement"]["nightPoolMax"] if v["entitlement"] else None, fmt=NUM)
r = tg(r, "Entitlement begins", lambda v: v["entitlement"]["begins"] if v["entitlement"] else None,
       "SlowSpace says Jan 2028; the construction programme ends Aug 2027 — conflict C-09.")
r += 1
r = tg(r, "Ordinary resolution (bps)", lambda v: v["governance"]["ordinaryBps"] if v["governance"] else None,
       "5001, not 5000. A tie is not approval.", fmt=NUM)
r = tg(r, "Special resolution (bps)", lambda v: v["governance"]["specialBps"] if v["governance"] else None, fmt=NUM)
r = tg(r, "Quorum (bps)", lambda v: v["governance"]["quorumBps"] if v["governance"] else None, fmt=NUM)
r = tg(r, "Reserved matters", lambda v: v["governance"]["reservedMatters"] if v["governance"] else None)
r = tg(r, "Transfer rule", lambda v: v["governance"]["transferRule"] if v["governance"] else None)
r = tg(r, "Designated partners", lambda v: v["governance"]["designatedPartners"] if v["governance"] else None)
r += 1
r = tg(r, "Minimum investment (bps)", lambda v: v["ladder"]["minimumInvestmentBps"], fmt=NUM)
r = tg(r, "Minimum unit / step (bps)", lambda v: v["ladder"]["minUnitBps"], fmt=NUM)
r = tg(r, "Ceiling (bps)", lambda v: v["ladder"]["ceilingBps"],
       "A governance limit, not a commercial one — above 50% one partner carries every ordinary "
       "resolution alone.", fmt=NUM)

r += 1
ws.cell(row=r, column=1,
        value="Amber cells: the intake states nothing. Solace has no entitlement and no governance row at all.").font = DIM


# ───────────────────────── 6 · ESTATE & DESIGN ─────────────────────────
ws = wb.create_sheet("Estate & Design")
head(ws, "THE GROUND, AND WHAT IS BUILT ON IT",
     "From the spatial ledger. Five estates, one building system, differing only by climate pack.", 11)
r = header_row(ws, 4, [
    "Estate", "Brand", "Region", "Climate pack", "Site (ac)", "Buildable (ac)",
    "Landscape held", "Keys", "Status", "Vehicle", "Character",
], [22, 11, 18, 22, 10, 12, 13, 7, 13, 14, 40])

e_start = r
for e in E:
    r = row(ws, r, [
        e["name"], e["brand"], e["region"], e["pack"], e["siteArea"], e["buildableEnvelope"],
        None, e["keys"], e["status"], e["vehicle"] or "—", e["character"],
    ], font=IN)
    L = r - 1
    c = ws.cell(row=L, column=7, value=f"=IF(E{L}=0,\"\",(E{L}-F{L})/E{L})")
    c.font = CALC; c.number_format = PCT; c.border = BOX
    if e["breaches"]:
        for col in (1, 8):
            ws.cell(row=L, column=col).fill = STOP
e_end = r - 1

r += 1
ws.cell(row=r, column=1, value="TOTAL KEYS").font = LBL
t = ws.cell(row=r, column=8, value=f"=SUM(H{e_start}:H{e_end})")
t.font = CALC; t.number_format = NUM; t.border = BOX
r += 2

ws.cell(row=r, column=1, value="WHERE AN ESTATE BREAKS THE PLATFORM'S OWN STANDARD").font = LBL
r += 1
any_breach = False
for e in E:
    for b in e["breaches"]:
        ws.cell(row=r, column=1, value=e["name"]).font = BAD
        c = ws.cell(row=r, column=2, value=b); c.font = TXT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
        any_breach = True
        r += 1
if not any_breach:
    ws.cell(row=r, column=1, value="None.").font = TXT
    r += 1
r += 1
ws.cell(row=r, column=1, value=D["landscapeNote"]).font = DIM
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
ws.row_dimensions[r].height = 30
r += 2

ws.cell(row=r, column=1, value="KEY TYPOLOGIES").font = LBL
r += 1
r = header_row(ws, r, ["Estate", "Typology", "Count", "Sqft each", "", "", "", "", "", "", "Note"])
for e in E:
    for k in e["keyTypes"]:
        r = row(ws, r, [e["name"], k["name"], k["count"], k["area"], None, None, None, None, None, None, k["note"]],
                font=IN, fmt=NUM)
        if k["area"] > 550:
            ws.cell(row=r - 1, column=4).font = BAD
r += 1

ws.cell(row=r, column=1, value="FOOTPRINT (sqft)").font = LBL
r += 1
r = header_row(ws, r, ["Estate", "Lodging built", "Working built", "Hardscape", "Softscape",
                       "", "", "", "", "", "Note"])
for e in E:
    f = e["footprint"]
    if not f: continue
    r = row(ws, r, [e["name"], f["lodgingBuilt"], f["workingBuilt"], f["hardscape"], f["softscape"],
                    None, None, None, None, None,
                    "The working half is meant to be invisible; a single built-area figure would hide "
                    "whether that held."], font=IN, fmt=NUM)
r += 1

ws.cell(row=r, column=1, value="ARCHITECTURAL LANGUAGE — common to every climate pack").font = LBL
r += 1
for a in D["language"]:
    ws.cell(row=r, column=1, value="·"); ws.cell(row=r, column=1).font = TXT
    ws.cell(row=r, column=2, value=a).font = TXT
    r += 1
r += 1

ws.cell(row=r, column=1, value="PLATFORM FINANCIAL CONSTITUTION").font = LBL
r += 1
for c_ in D["constitution"]:
    ws.cell(row=r, column=1, value=c_["standard"]).font = TXT
    ws.cell(row=r, column=2, value=c_["rule"]).font = IN
    r += 1


# ───────────────────────── 7 · PIPELINE ─────────────────────────
ws = wb.create_sheet("Pipeline")
head(ws, "THE STANDARDISED PIPELINE",
     "Ten lifecycle stages every vehicle passes through, and the construction programme from the ledger.", 6)
r = header_row(ws, 4, ["Stage", "Name", "Purpose", "Gate", "Owner", "Where the three are"],
               [8, 20, 56, 20, 24, 34])

for s in S:
    at = []
    for v in V:
        if s["id"] == "01" and v["lifecycle"] == "forming": at.append(v["key"])
        elif s["id"] == "02" and v["lifecycle"] == "raising": at.append(v["key"])
    r = row(ws, r, [s["id"], s["label"], s["purpose"], s["gate"], s["owner"],
                    ", ".join(at) if at else ""], font=TXT)
    if at:
        for col in range(1, 7):
            ws.cell(row=r - 1, column=col).fill = BAND

r += 2
ws.cell(row=r, column=1, value="CONSTRUCTION PROGRAMME — from the deployment gantt").font = LBL
r += 1
r = header_row(ws, r, ["Estate", "Keys", "Start", "End", "Vehicle", "Note"])
for e in E:
    p = e["programme"]
    r = row(ws, r, [e["name"], e["keys"], p["start"] if p else "", p["end"] if p else "",
                    e["vehicle"] or "—", e["strategicRole"]], font=IN)

r += 2
ws.cell(row=r, column=1,
        value="Every project is locked to twelve months and a single Mivan formwork set moves between "
              "sites, which is why the starts are staggered rather than parallel.").font = DIM
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)


# ───────────────────────── 8 · CONFLICTS ─────────────────────────
ws = wb.create_sheet("Conflicts")
head(ws, "WHAT DOES NOT RECONCILE",
     "Nine contradictions across the intake, the spatial ledger and the modelled canon. Read this before "
     "acting on any figure in this workbook.", 6)
r = header_row(ws, 4, ["ID", "Vehicle", "Severity", "What", "The two sides", "Who settles it"],
               [7, 13, 12, 40, 62, 34])

for c_ in C:
    blocking = c_["severity"] == "blocking"
    r = row(ws, r, [
        c_["id"], c_["vehicle"], c_["severity"].upper(), c_["what"],
        "\n".join("• " + s for s in c_["sides"]) + "\n\n" + c_["why"],
        c_["settledBy"],
    ], font=TXT, fill=STOP if blocking else WARN)
    ws.cell(row=r - 1, column=3).font = BAD if blocking else LBL
    ws.row_dimensions[r - 1].height = 84

r += 1
n_block = sum(1 for c_ in C if c_["severity"] == "blocking")
ws.cell(row=r, column=1,
        value=f"{n_block} of {len(C)} are blocking. A blocking conflict stops its vehicle reaching a public "
              f"surface — that gate is a function in constants/vehicles.ts, not a convention, because a "
              f"convention is what gets forgotten under a launch date.").font = DIM
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

os.makedirs("outputs", exist_ok=True)
out = "outputs/GC-COLLECTION-FULL-VIEW.xlsx"
wb.save(out)
print("wrote", out)
