#!/usr/bin/env python3
"""
Wave 1 Deliverables Tracker — Foundation Lock
Organized by category with status tracking
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "W1 Deliverables"

# ── Headers ──────────────────────────────────────────────────────────
headers = [
    "Deliverable",
    "Category",
    "Status",
    "Source/Location",
    "Notes",
    "Owner",
    "Due",
    "Priority"
]

ws.append(headers)
header_row = ws[1]

# Style headers
header_fill = PatternFill(start_color="0C3024", end_color="0C3024", fill_type="solid")  # forest
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

for cell in header_row:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Data ─────────────────────────────────────────────────────────────
deliverables = [
    # Content & Copy (orange priority)
    ("Enterprise Constitution (L1 document)", "Content", "TBD", "GC ALL LAYERS OF DATABASE", "Full immutable principles text; Non-negotiable values; Constitutional hierarchy; Constitutional laws", "", "", "HIGH"),
    ("Constitutional Thesis", "Content", "TBD", "PRD Vol I", "What GC is; What GC is not; Three constitutional dimensions; Constitutional success/failure criteria", "", "", "HIGH"),
    ("Brand Constitution (L1 sub-doc)", "Content", "TBD", "Brand docs (external)", "Brand voice and tone; Brand promises; Brand prohibitions", "", "", "MEDIUM"),
    ("Enterprise Vocabulary", "Content", "PARTIAL", "GC ALL LAYERS; Design System", "Complete forbidden/replaced terms list (Room→Studio, Customer→Guest, Booking→Journey, Housekeeping→Studio Care, etc.); Preferred terminology glossary (50–100 terms); Context for each term", "", "", "HIGH"),

    # Data & Schema
    ("Business Object Taxonomy (closed list)", "Data", "HAVE", "GC BUSINESS OBJECTS", "BO-01…BO-12 enumerated and classified", "", "", "HIGH"),
    ("Business Object Descriptions", "Data", "HAVE", "GC BUSINESS OBJECTS + GC ELEMENTS", "12 objects with mission, classification, field hints", "", "", "HIGH"),

    # Business Rules
    ("Founding Invariants Register", "Rules", "PARTIAL", "GC ALL LAYERS + WAVE-INTAKE.md", "Studio requires Property; Ledger append-only; Knowledge immutable; Every capability publishes events; Every decision has provenance; [+customer-specific invariants]", "", "", "HIGH"),
    ("Enterprise Policies (summary)", "Rules", "TBD", "Brand/Legal docs (external)", "Privacy, Security, Accessibility, Sustainability, Ethics, Luxury standard (for system design impact)", "", "", "HIGH"),

    # Visual Assets
    ("Design System Package (v3.0)", "Assets", "HAVE", "GC-DesignSystem-Canonical (2).html", "Colour ontology, Typography, Spacing (4px base), Motion, IL-1…IL-6, Density modes, Visual modes", "", "", "HIGH"),
    ("Token Package (machine-readable)", "Assets", "TBD", "Extract from HTML", "CSS custom properties file or JSON; Versioned immutably (v3.0 locked); One source, consumed by all layers", "", "", "HIGH"),

    # Configuration
    ("Vocabulary Linter Configuration", "Config", "TBD", "Create ESLint/TypeScript rule", "Forbidden terms list as rule; CI integration point (should fail build on violation)", "", "", "MEDIUM"),
    ("Build System Setup", "Config", "TBD", "Next.js project structure", "package.json, tsconfig, .eslintrc, vitest/jest config", "", "", "MEDIUM"),

    # Test Data
    ("Sample Objects (one per BO-01…12)", "Test", "TBD", "Create JSON fixtures", "Use real or realistic values; Include edge cases (optional null fields, min/max values); Include one error case per object", "", "", "LOW"),

    # Integration
    ("Project Structure & Git Setup", "Integration", "TBD", "C:\\gc-app", "Initialize Next.js project, Git, folder structure", "", "", "HIGH"),
]

for row_idx, (deliverable, category, status, source, notes, owner, due, priority) in enumerate(deliverables, start=2):
    ws.append([deliverable, category, status, source, notes, owner, due, priority])

    # Apply conditional formatting based on status
    status_cell = ws[f"C{row_idx}"]
    if status == "HAVE":
        status_cell.fill = PatternFill(start_color="1FAA59", end_color="1FAA59", fill_type="solid")  # confirm
        status_cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    elif status == "PARTIAL":
        status_cell.fill = PatternFill(start_color="E8672E", end_color="E8672E", fill_type="solid")  # hazard
        status_cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    elif status == "TBD":
        status_cell.fill = PatternFill(start_color="E8E8E6", end_color="E8E8E6", fill_type="solid")  # mist
        status_cell.font = Font(name="Arial", size=10, color="000000")

    # Priority color
    priority_cell = ws[f"H{row_idx}"]
    if priority == "HIGH":
        priority_cell.font = Font(name="Arial", size=10, bold=True, color="FF3B30")  # critical
    elif priority == "MEDIUM":
        priority_cell.font = Font(name="Arial", size=10, color="E8672E")  # hazard
    else:
        priority_cell.font = Font(name="Arial", size=10, color="6B6B6B")  # steel

    # Wrap text in notes
    ws[f"E{row_idx}"].alignment = Alignment(wrap_text=True, vertical="top")

    # Borders
    for col in range(1, 9):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = Border(
            left=Side(style="thin", color="6B6B6B"),
            right=Side(style="thin", color="6B6B6B"),
            top=Side(style="thin", color="6B6B6B"),
            bottom=Side(style="thin", color="6B6B6B")
        )

# ── Column widths ────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 45
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 35
ws.column_dimensions["E"].width = 50
ws.column_dimensions["F"].width = 15
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 10

# ── Row heights ──────────────────────────────────────────────────────
ws.row_dimensions[1].height = 25
for row in range(2, len(deliverables) + 2):
    ws.row_dimensions[row].height = 40

# ── Summary section ──────────────────────────────────────────────────
summary_row = len(deliverables) + 4
ws[f"A{summary_row}"] = "SUMMARY"
ws[f"A{summary_row}"].font = Font(name="Arial", size=12, bold=True)

summary_row += 1
ws[f"A{summary_row}"] = "Status Breakdown:"
ws[f"A{summary_row}"].font = Font(name="Arial", size=10, bold=True)

summary_row += 1
ws[f"A{summary_row}"] = "[HAVE] Ready to use:"
ws[f"B{summary_row}"] = "=COUNTIF(C2:C14,\"HAVE\")"
ws[f"B{summary_row}"].font = Font(name="Arial", size=10, bold=True, color="1FAA59")

summary_row += 1
ws[f"A{summary_row}"] = "[PARTIAL] Needs completion:"
ws[f"B{summary_row}"] = "=COUNTIF(C2:C14,\"PARTIAL\")"
ws[f"B{summary_row}"].font = Font(name="Arial", size=10, bold=True, color="E8672E")

summary_row += 1
ws[f"A{summary_row}"] = "[TBD] Not started:"
ws[f"B{summary_row}"] = "=COUNTIF(C2:C14,\"TBD\")"
ws[f"B{summary_row}"].font = Font(name="Arial", size=10, bold=True)

summary_row += 2
ws[f"A{summary_row}"] = "Total Deliverables:"
ws[f"B{summary_row}"] = len(deliverables)
ws[f"B{summary_row}"].font = Font(name="Arial", size=10, bold=True)

# ── Save ─────────────────────────────────────────────────────────────
wb.save("WAVE-1-DELIVERABLES.xlsx")
print("[OK] Created WAVE-1-DELIVERABLES.xlsx")
print(f"  {len(deliverables)} deliverables tracked")
print("  Open and fill in Owner + Due columns, then save")
